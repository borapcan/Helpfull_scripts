import openpyxl

# Load the two Excel files to compare
file1 = openpyxl.load_workbook('New all together.xlsx')
file2 = openpyxl.load_workbook('All together.xlsx')

# Get the active sheet for each file
sheet1 = file1.active
sheet2 = file2.active

# Create an empty list to store the differences
differences = []

# Iterate over each cell in the sheets and compare their values
for row in range(1, sheet1.max_row + 1):
    for col in range(1, sheet1.max_column + 1):
        cell1 = sheet1.cell(row=row, column=col)
        cell2 = sheet2.cell(row=row, column=col)
        if cell1.value != cell2.value:
            # Set the fill color of the cell to red
            cell1.fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            # Add the difference to the list
            differences.append(f"Row {row}, Column {col}: {cell1.value} != {cell2.value}")

# Save the modified file
file1.save('file1_modified.xlsx')

# Write the differences to a text file
with open('differences.txt', 'w') as file:
    if differences:
        file.write("Differences found:\n")
        file.write('\n'.join(differences))
    else:
        file.write("No differences found.")

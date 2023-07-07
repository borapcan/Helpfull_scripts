import random 
import openpyxl

# create a list of samples, blanks and standards for the user
plate_list = [f"stand_{i}_{j}" for i in range(4, 7) for j in range(1, 5)]
plate_list += [f"blank_{i}" for i in range(1, 5)]
plate_list += [f"sample_{i}" for i in range(1, 81)]

# shuffle 
random.shuffle(plate_list)

with open('sample_list.csv', 'w') as f:
    for value in plate_list:
        f.write(str(value) + '\n')

# create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# write the plates column names
worksheet.cell(row=1, column=1).value = ''
for col in range(2, 14):
    worksheet.cell(row=1, column=col).value = col-1

# write the plates row names
for row in range(2, 10):
    worksheet.cell(row=row, column=1).value = chr(row+63)
    for col in range(2, 15):
        worksheet.cell(row=row, column=col).value = ''

col_idx = 2
for j in range(0, 12):
    for i in range(1, 9):
        cell = worksheet.cell(row=i+1, column=col_idx)
        cell.value = plate_list[i+j*8-1]
        if cell.value.startswith("stand_4"):
            cell.fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        elif cell.value.startswith("stand_5"):
            cell.fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        elif cell.value.startswith("stand_6"):
            cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        elif cell.value.startswith("blank"):
            cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    col_idx += 1



# save the Excel file
workbook.save('output.xlsx')
import random
import openpyxl

# create a list of samples, blanks and standards for the user
sample_list = ['1.1', '2.1', '2.2', '2.3', '3.1', '3.2', '3.3', '4.1', '4.2', '4.3', '5.1', '6.1', '6.3', '7.1', '7.2', '8.1', '8.3', '9.1', '9.2', '9.3', '10.1', '10.2', '10.3', '11.1', '12.1', '12.2', '12.3', '13.1', '13.2', '13.3', '14.1', '14.2', '14.3', '15.1', '16.1', '16.2', '16.3', '17.1', '17.2', '17.3', '18.1', '18.2', '18.3', '19.1', '19.3', '20.1', '20.2', '21.1', '22.1', '23.1', '23.2', '24.1', '24.3', '25.1', '25.2', '26.1', '27.1', '28.1', '28.3', '29.1', '29.2', '30.1', '31.1', '32.1', '32.3', '33.1', '34.1', '35.1', '35.3', '36.1', '37.1', '37.2', '37.3', '38.1', '39.1', '40.1', '40.2', '40.3', '41.1', '41.2', '42.1', '43.1', '43.2', '43.3', '44.1', '45.1', '46.1', '46.3', '47.1', '47.3', '48.1', '48.2', '48.3', '49.1', '50.1', '51.1', '52.1', '52.2', '52.3', '53.1', '53.2', '53.3', '54.1', '54.3', '55.1', '55.2', '55.3', '56.1', '57.1', '57.3', '58.1', '59.1', '59.3', '60.1', '60.2', '60.3', '61.1', '62.1', '62.3', '63.1', '63.3', '64.1', '65.1', '65.3', '66.1', '66.3', '67.1', '67.3', '69.1', '70.1']

# shuffle the list
random.shuffle(sample_list)

# divide the list into two smaller lists
half_len = len(sample_list) // 2
plate_list1 = sample_list[:half_len]
plate_list2 = sample_list[half_len:]

plate_list1 += [f"stand_{i}_{j}" for i in range(4, 7) for j in range(1, 5)]
plate_list1 += [f"blank_{i}" for i in range(1, 5)]
plate_list2 += [f"stand_{i}_{j}" for i in range(4, 7) for j in range(1, 5)]
plate_list2 += [f"blank_{i}" for i in range(1, 5)]

# shuffle 
random.shuffle(plate_list1)
random.shuffle(plate_list2)

desired_length = 96

# add empty strings to plate_list1
plate_list1 += [""] * (desired_length - len(plate_list1))

# add empty strings to plate_list2
plate_list2 += [""] * (desired_length - len(plate_list2))

with open('sample_list1.csv', 'w') as f:
    for value in plate_list1:
        f.write(str(value) + '\n')

with open('sample_list2.csv', 'w') as f:
    for value in plate_list2:
        f.write(str(value) + '\n')

# create a new Excel workbook and worksheet
workbook1 = openpyxl.Workbook()
worksheet1 = workbook1.create_sheet('Plate 1')

# write the plates column names
worksheet1.cell(row=1, column=1).value = ''
for col in range(2, 14):
    worksheet1.cell(row=1, column=col).value = col-1

# write the plates row names
for row in range(2, 10):
    worksheet1.cell(row=row, column=1).value = chr(row+63)
    for col in range(2, 15):
        worksheet1.cell(row=row, column=col).value = ''

col_idx = 2
for j in range(0, 12):
    for i in range(1, 9):
        cell = worksheet1.cell(row=i+1, column=col_idx)
        cell.value = plate_list1[i+j*8-1]
        if cell.value.startswith("stand_4"):
            cell.fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        elif cell.value.startswith("stand_5"):
            cell.fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        elif cell.value.startswith("stand_6"):
            cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        elif cell.value.startswith("blank"):
            cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    col_idx += 1

worksheet2 = workbook1.create_sheet('Plate 2')

# write the plates column names
worksheet2.cell(row=1, column=1).value = ''
for col in range(2, 14):
    worksheet2.cell(row=1, column=col).value = col-1

# write the plates row names
for row in range(2, 10):
    worksheet2.cell(row=row, column=1).value = chr(row+63)
    for col in range(2, 15):
        worksheet2.cell(row=row, column=col).value = ''

col_idx = 2
for j in range(0, 12):
    for i in range(1, 9):
        cell = worksheet2.cell(row=i+1, column=col_idx)
        cell.value = plate_list2[i+j*8-1]
        if cell.value.startswith("stand_4"):
            cell.fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        elif cell.value.startswith("stand_5"):
            cell.fill = openpyxl.styles.PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        elif cell.value.startswith("stand_6"):
            cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        elif cell.value.startswith("blank"):
            cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    col_idx += 1



# save the Excel file
workbook1.save('output.xlsx')